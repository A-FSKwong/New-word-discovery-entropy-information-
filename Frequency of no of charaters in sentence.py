import sys
import subprocess

def install_and_import(module_name, package_name=None):
    """
    檢查指定的模組是否已安裝，若未安裝則自動安裝指定的套件，再匯入該模組。
    module_name: 匯入時使用的模組名稱
    package_name: pip 安裝時使用的套件名稱，若未指定則預設與 module_name 相同
    """
    if package_name is None:
        package_name = module_name
    try:
        __import__(module_name)
    except ImportError:
        print(f"未偵測到模組 '{module_name}'，開始安裝套件 '{package_name}'...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
        print(f"套件 '{package_name}' 安裝完成。")
        __import__(module_name)

# 自動安裝所需模組
install_and_import("docx", "python-docx")
install_and_import("pandas")
install_and_import("openpyxl")

import os
import re
import docx
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import AreaChart, Reference

def read_docx(file_path):
    """
    讀取 DOCX 檔案並回傳所有文字。
    """
    doc = docx.Document(file_path)
    full_text = ""
    for para in doc.paragraphs:
        full_text += para.text
    return full_text

def split_sentences(text):
    """
    利用正則表達式依標點符號分割句子。
    此處定義的標點符號包含：中文標點（，。！？、；：）、英文標點（,.?!;:）以及空白字元，
    使得被空白分隔的文字也視為一句話。
    """
    pattern = r'[\s，。！？、；：,.?!;:]+'  
    sentences = re.split(pattern, text)
    # 移除空白或空字串
    sentences = [s.strip() for s in sentences if s.strip() != ""]
    return sentences

def count_sentence_lengths(sentences):
    """
    計算每個句子的字數，並統計句子長度（介於 2 到 50）的出現次數。
    """
    # 初始化字典，key 為句子長度（2~50），value 為出現次數
    length_count = {i: 0 for i in range(2, 51)}
    for sentence in sentences:
        length = len(sentence)
        if 2 <= length <= 50:
            length_count[length] += 1
    return length_count

def main():
    # 提示使用者輸入 DOCX 檔案路徑
    input_path = input("請輸入 DOCX 檔案的完整路徑 (例如: c:/a.docx): ").strip()
    if not os.path.isfile(input_path):
        print("檔案不存在，請確認路徑是否正確。")
        return

    try:
        text = read_docx(input_path)
    except Exception as e:
        print(f"讀取 DOCX 檔案時發生錯誤：{e}")
        return

    # 分割句子，統計句子長度（2 ~ 50）
    sentences = split_sentences(text)
    total_sentences = len(sentences)  # 總句數
    length_count = count_sentence_lengths(sentences)

    # 準備輸出資料，建立三個欄位：
    # "No. of Charater"、"Frequency"、以及 "%"（數值以兩位小數，不附%符號）
    numbers = []
    frequencies = []
    percentages = []
    for num in range(2, 51):
        numbers.append(num)
        freq = length_count[num]
        frequencies.append(freq)
        percent = (freq / total_sentences * 100) if total_sentences > 0 else 0
        percentages.append(round(percent, 2))

    df = pd.DataFrame({
        "No. of Charater": numbers,
        "Frequency": frequencies,
        "%": percentages
    })

    # 設定輸出 Excel 檔案路徑（與輸入檔案相同目錄，檔名為 output.xlsx）
    output_dir = os.path.dirname(input_path)
    output_path = os.path.join(output_dir, "output.xlsx")

    try:
        df.to_excel(output_path, index=False)
        print(f"Excel 檔案已成功儲存到：{output_path}")
    except Exception as e:
        print(f"儲存 Excel 檔案時發生錯誤：{e}")
        return

    try:
        # 利用 openpyxl 載入 Excel 檔案
        wb = load_workbook(output_path)
        ws = wb.active

        # 新增一個工作表 "ChartData"，將原資料中僅保留「Frequency」值小於或等於 30 的資料列，
        # 並複製「No. of Charater」及「%」兩欄，供圖表使用
        chart_ws = wb.create_sheet("ChartData")
        # 複製標題列：取「No. of Charater」及「%」兩欄
        chart_ws.append([ws.cell(row=1, column=1).value, ws.cell(row=1, column=3).value])
        # 複製資料列（篩選 Frequency <= 30）
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3, values_only=True):
            # row[0]: No. of Charater, row[1]: Frequency, row[2]: %
            if row[1] <= 30:
                chart_ws.append([row[0], row[2]])

        # 若 ChartData 工作表僅有標題列，則不建立圖表
        if chart_ws.max_row < 2:
            print("篩選後無任何 Frequency 資料（頻率皆大於 30），故不建立圖表。")
        else:
            # 建立 Area (面積) 圖表，資料來源採用 ChartData 工作表中第二欄（%數值）
            chart = AreaChart()
            chart.style = 10
            chart.title = "句子字數百分比統計圖"
            chart.y_axis.title = "Percentage"
            chart.x_axis.title = "No. of Charater"

            # 設定圖表資料來源：ChartData 工作表中第2欄資料（包含標題）
            data = Reference(chart_ws, min_col=2, min_row=1, max_row=chart_ws.max_row, max_col=2)
            chart.add_data(data, titles_from_data=True)
            # 設定 X 軸類別：ChartData 工作表中第1欄（不含標題）
            cats = Reference(chart_ws, min_col=1, min_row=2, max_row=chart_ws.max_row)
            chart.set_categories(cats)
            # 將圖表插入至原工作表指定位置 (例如：D2 儲存格起始)
            ws.add_chart(chart, "D2")
            print("圖表已成功新增至 Excel 檔案中。")

        wb.save(output_path)
    except Exception as e:
        print(f"新增圖表時發生錯誤：{e}")

if __name__ == "__main__":
    main()
